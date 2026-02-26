import os
from typing import List
from openpyxl import load_workbook
from .base import BaseExtractor, ContentItem


class XLSXExtractor(BaseExtractor):
    """
    Extracts text content from Excel (.xlsx) files, row by row.
    
    Each data row becomes its own ContentItem with column headers baked in,
    so downstream embeddings retain full context for tabular data.
    """

    def extract(self, file_path: str) -> List[ContentItem]:
        items = []
        wb = load_workbook(file_path, read_only=True, data_only=True)
        filename = os.path.basename(file_path)

        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            if len(all_rows) < 2:
                # Only header or empty — store as-is
                if all_rows:
                    cell_values = [str(c) for c in all_rows[0] if c is not None]
                    if cell_values:
                        items.append(ContentItem(
                            content=f"[Sheet: {sheet_name}] " + " | ".join(cell_values),
                            type="text",
                            source=filename,
                            page_num=sheet_index + 1,
                            metadata={"sheet_name": sheet_name}
                        ))
                continue

            # First row is the header
            headers = [str(c) if c is not None else "" for c in all_rows[0]]

            for row_index, row in enumerate(all_rows[1:], start=1):
                pairs = []
                for col, val in zip(headers, row):
                    if val is not None:
                        val_str = str(val).strip()
                        if val_str:
                            pairs.append(f"{col}: {val_str}")

                if pairs:
                    text = f"[Sheet: {sheet_name}] " + ", ".join(pairs)
                    items.append(ContentItem(
                        content=text,
                        type="text",
                        source=filename,
                        page_num=sheet_index + 1,
                        metadata={"sheet_name": sheet_name, "row_index": row_index}
                    ))

        wb.close()
        return items
